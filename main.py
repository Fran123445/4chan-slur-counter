import requests
import bs4
import time
import multiprocessing
from openpyxl import workbook
from openpyxl.styles import Alignment


def get_threads(board_dict, board):
    """Get all active (non-archived) threads on a given board"""

    page = 0
    thread_list = []
    page_list = []

    # Do a request for each page on each board
    while page < 11:
        page_list.append(requests.get(f"https://boards.4chan.org/{board}/{page}"))
        if page == 0:
            page += 2
        else:
            page += 1

    page = 0

    # Get each thread number
    while page < 10:
        soup = bs4.BeautifulSoup(page_list[page].text, "html.parser")
        for tag in soup.findAll("span", class_="postNum desktop"):
            thread_list.append(str(tag).split("#")[0].split("/")[1])
            thread_list = list(dict.fromkeys(thread_list))
        page += 1

    print(f"/{board}/ - Requesting done")
    board_dict[board] = thread_list


def count(board_dict, board):
    """Count the amount of slurs per board"""

    slur_list = {"nigga": 0, "nigger": 0, "fag": 0, "troon": 0, "tranny": 0, "(((they)))": 0, "(((them)))": 0, "kike": 0,
                 "argie": 0, "bri'ish": 0, "dyke": 0, "chink": 0, "pajeet": 0, "goy": 0, "gypsy": 0, "tard": 0, "schizo": 0,
                 "total amount of posts": 0, "total amount of slurs": 0, "slurs per post": 0}

    for thread in board_dict[board]:
        # Iterate through each thread on the current board

        request = requests.get(f"https://boards.4chan.org/{board}/thread/{thread}")
        soup = bs4.BeautifulSoup(request.text, "html.parser")
        for tag in soup.findAll("blockquote", class_="postMessage"):
            # Iterate through each reply on the current thread

            tag_list = str(tag).split()
            for word in tag_list:
                # Iterate through each word on the current reply

                if any(slur in word.lower() for slur in slur_list):
                    # Check if any slur is contained in any of those words

                    for slur in slur_list:
                        if slur in word.lower():
                            slur_list[slur] += 1
                            break

            slur_list["total amount of posts"] += 1

    board_dict[board] = slur_list

    print(f"/{board}/ - Counting done")

    return board_dict


def multiproc(board_dict, func):

    processes = []
    manager = multiprocessing.Manager()
    aux = manager.dict(board_dict)

    for board in board_dict:
        get_pages = multiprocessing.Process(target=func, args=(aux, board))
        processes.append(get_pages)
        get_pages.start()

    for proc in processes:
        proc.join()

    return aux


def calc_slur_per_post(board_dict):
    """Calculate the amount of slurs per post on average"""

    for board in board_dict:
        slur_sum = 0
        aux = board_dict[board]

        for key in list(board_dict[board].keys())[:-3]:
            slur_sum += board_dict[board][key]

        try:
            slurs_per_post = slur_sum/board_dict[board]["total amount of posts"]
        except ZeroDivisionError:
            slurs_per_post = 0

        slurs_per_post = round(slurs_per_post, 6)

        aux["total amount of slurs"] = slur_sum
        aux["slurs per post"] = slurs_per_post

        board_dict[board] = aux

    return board_dict


def count_to_txt(board_slur_list, current_time):
    """Log each board's slur count on a file"""

    time = f"{current_time.tm_hour}:{current_time.tm_min}\t{current_time.tm_mday}/{current_time.tm_mon}/{current_time.tm_year}"

    with open("slur_count.txt", "a") as log:
        log.write(f"{time}\n\n")

        for board in board_slur_list:
            log.write(f"/{board}/\n")
            for slur in board_slur_list[board]:
                log.write(f'"{slur}": {board_slur_list[board][slur]}\t')
            log.write("\n\n")

        log.write("\n\n\n")

    log.close()

    print("Txt file done.")


def count_to_xlsx(board_dict):

    wb = workbook.Workbook()
    ws = wb.active

    ws.freeze_panes = "B2"
    
    current_column = 2

    for slur in board_dict[board_dict.keys()[0]]:
        cell = ws.cell(row=1, column=current_column, value=slur)

        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = len(slur)+2
        current_column += 1

    current_column = 1
    current_row = 2

    for board in board_dict:

        cell_value = ws.cell(row=current_row, column=current_column, value=f"/{board}/")

        current_column += 1

        for amount in board_dict[board].values():
            cell_value = ws.cell(row=current_row, column=current_column, value=amount)
            current_column += 1

        current_column = 1
        current_row += 1

    wb.save('slur_count.xlsx')
    print("Xlsx file done.")


if __name__ == '__main__':
    boards = ["3", "a", "adv", "an", "b", "bant", "biz", "c", "cgl", "ck", "cm", "co", "d", "diy",
              "e", "f", "fa", "fit", "g", "gd", "gif", "h", "hc", "hm", "hr", "i", "ic", "int", "jp",
              "k", "lgbt", "lit", "m", "mlp", "mu", "n", "news", "o", "out", "p", "po", "pol", "r",
              "r9k", "s4s", "s", "sci", "soc", "sp", "t", "tg", "toy", "trv", "tv", "u", "v", "vg",
              "vp", "vr", "vt", "w", "wg", "wsg", "wsr", "x", "xs", "y"]

    board_dict = dict.fromkeys(boards)

    print("\n-Starting requesting process-\n")
    board_dict = multiproc(board_dict, get_threads)

    print("\n-Starting count process-\n")
    board_dict = multiproc(board_dict, count)

    current_time = time.localtime(time.time())
    board_dict = calc_slur_per_post(board_dict)
    count_to_txt(board_dict, current_time)
    count_to_xlsx(board_dict)
